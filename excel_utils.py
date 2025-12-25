from openpyxl import load_workbook
from datetime import datetime

EXCEL_FILE = "excel/invite_balance.xlsx"
SHEET_NAME = "balances"

class ExcelWriter:
    """
    Batch Excel writer.
    Opens the workbook once, updates multiple rows, and saves once.
    """

    def __init__(self):
        self.wb = load_workbook(EXCEL_FILE)
        self.ws = self.wb[SHEET_NAME]
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Map: site_name -> row index
        self.row_map = {}
        for row in range(2, self.ws.max_row + 1):
            name = self.ws.cell(row, 1).value
            if name:
                self.row_map[name] = row

    def update(self, site_name: str, balance: float):
        if site_name not in self.row_map:
            raise ValueError(f"Site not found in Excel: {site_name}")

        row = self.row_map[site_name]
        self.ws.cell(row, 2).value = balance
        self.ws.cell(row, 3).value = self.timestamp

    def save(self):
        self.wb.save(EXCEL_FILE)
